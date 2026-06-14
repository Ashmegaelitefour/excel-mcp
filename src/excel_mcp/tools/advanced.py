"""Advanced tools: join_sheets, pivot_data, unpivot_data, concat_sheets,
window_function, save_sheet, clone_sheet, export_data, run_query."""

from __future__ import annotations

from typing import Literal, Optional

import polars as pl

from excel_mcp import state
from excel_mcp.security import validate_path
from excel_mcp.tools.load import _load, _lazy, _df_to_markdown


def join_sheets(
    left_file: str,
    left_sheet: str,
    right_file: str,
    right_sheet: str,
    on: list[str],
    how: Literal["inner", "left", "right", "full", "cross"] = "inner",
) -> str:
    """Join two sheets/files on one or more key columns.

    Both files can be any supported format (.xlsx, .xls, .csv, .tsv, .parquet).
    The joined result is stored in the left sheet's cache slot.

    Args:
        left_file: Path to the left file.
        left_sheet: Sheet name in the left file.
        right_file: Path to the right file (may be the same as left_file).
        right_sheet: Sheet name in the right file.
        on: List of column names to join on (must exist in both sheets).
        how: Join type — inner, left, right, full, or cross.
    """
    left_lf = _lazy(left_file, left_sheet)
    right_lf = _lazy(right_file, right_sheet)

    # Resolve column name conflicts on right side before joining
    right_cols = _load(right_file, right_sheet).columns
    left_cols = _load(left_file, left_sheet).columns
    right_non_key = [c for c in right_cols if c not in on]
    rename = {c: f"{c}_right" for c in right_non_key if c in left_cols}
    if rename:
        right_lf = right_lf.rename(rename)

    joined = left_lf.join(right_lf, on=on, how=how).collect()
    abs_left = validate_path(left_file)
    state.set_df(abs_left, left_sheet, joined)
    return (
        f"Joined shape: {joined.height} rows × {joined.width} columns\n\n"
        + _df_to_markdown(joined.head(20))
    )


def pivot_data(
    file_path: str,
    sheet_name: str,
    index: list[str],
    on: str,
    values: str,
    aggregate_fn: Literal["sum", "mean", "count", "min", "max", "first", "last"] = "sum",
) -> str:
    """Pivot a sheet: turn unique values of one column into new column headers.

    Args:
        file_path: Path to the file.
        sheet_name: Sheet name.
        index: Columns to use as row identifiers.
        on: Column whose unique values become the new column headers.
        values: Column whose values fill the pivot cells.
        aggregate_fn: Aggregation applied when multiple rows map to the same cell.
    """
    # pivot() requires a DataFrame — collect lazily with only needed columns first
    df = _lazy(file_path, sheet_name).select([*index, on, values]).collect()
    pivoted = df.pivot(values=values, index=index, on=on, aggregate_function=aggregate_fn)
    return (
        f"Pivoted: {pivoted.height} rows × {pivoted.width} columns\n\n"
        + _df_to_markdown(pivoted.head(30))
    )


def unpivot_data(
    file_path: str,
    sheet_name: str,
    id_vars: list[str],
    value_vars: Optional[list[str]] = None,
    variable_name: str = "variable",
    value_name: str = "value",
) -> str:
    """Unpivot (melt) wide-format data into long format.

    Args:
        file_path: Path to the file.
        sheet_name: Sheet name.
        id_vars: Columns to keep as identifier variables.
        value_vars: Columns to unpivot. Defaults to all non-id columns.
        variable_name: Name for the new variable column (default "variable").
        value_name: Name for the new value column (default "value").
    """
    # Select only the columns we need before collecting
    select_cols = id_vars + (value_vars or [c for c in _load(file_path, sheet_name).columns if c not in id_vars])
    df = _lazy(file_path, sheet_name).select(select_cols).collect()
    melted = df.unpivot(
        index=id_vars,
        on=value_vars,
        variable_name=variable_name,
        value_name=value_name,
    )
    return (
        f"Unpivoted: {melted.height} rows × {melted.width} columns\n\n"
        + _df_to_markdown(melted.head(30))
    )


def concat_sheets(
    sources: list[dict],
    how: Literal["vertical", "horizontal", "diagonal"] = "vertical",
) -> str:
    """Concatenate multiple sheets or files into one result.

    For vertical (row-wise) concat, all sources must have the same columns or use
    how="diagonal" to align by column name and fill missing values with null.

    Args:
        sources: List of dicts with "file" and "sheet" keys, e.g.
                 [{"file": "data.xlsx", "sheet": "Jan"},
                  {"file": "data.xlsx", "sheet": "Feb"}]
        how: Concatenation direction — "vertical" (stack rows), "horizontal"
             (side-by-side columns), or "diagonal" (vertical with column alignment).
    """
    frames = []
    for src in sources:
        file_path = src.get("file", "")
        sheet_name = src.get("sheet", "")
        if not file_path or not sheet_name:
            raise ValueError(f"Each source must have 'file' and 'sheet' keys. Got: {src}")
        frames.append(_lazy(file_path, sheet_name))

    # pl.concat on LazyFrames — Polars optimises across all sources before collecting
    result = pl.concat(frames, how=how).collect()
    return (
        f"Concatenated {len(frames)} sources ({how}): "
        f"{result.height} rows × {result.width} columns\n\n"
        + _df_to_markdown(result.head(20))
    )


def window_function(
    file_path: str,
    sheet_name: str,
    column: str,
    func: Literal["rank", "cumsum", "cumprod", "cummax", "cummin", "diff", "pct_change", "rolling_mean", "rolling_sum"],
    partition_by: Optional[list[str]] = None,
    order_by: Optional[list[str]] = None,
    window_size: int = 3,
    output_column: Optional[str] = None,
) -> str:
    """Apply a window / analytic function to a column and add it as a new column.

    Useful for running totals, rankings, rolling averages, and percent-change
    calculations — all without collapsing rows.

    Supported functions:
      rank, cumsum, cumprod, cummax, cummin, diff, pct_change,
      rolling_mean (requires window_size), rolling_sum (requires window_size).

    Args:
        file_path: Path to the file.
        sheet_name: Sheet name.
        column: Column to apply the function to.
        func: Window function name.
        partition_by: Optional list of columns to reset the window per group
                      (like SQL PARTITION BY).
        order_by: Columns to sort by within each window partition.
        window_size: Number of rows for rolling functions (default 3).
        output_column: Name for the new column. Defaults to "{column}_{func}".
    """
    existing_cols = _load(file_path, sheet_name).columns
    if column not in existing_cols:
        raise ValueError(f"Column '{column}' not found. Available: {existing_cols}")

    out_col = output_column or f"{column}_{func}"
    base = pl.col(column)
    lf = _lazy(file_path, sheet_name)

    if order_by:
        lf = lf.sort(order_by)

    rolling_funcs = {"rolling_mean", "rolling_sum"}
    if func in rolling_funcs:
        expr = base.rolling_mean(window_size=window_size) if func == "rolling_mean" else base.rolling_sum(window_size=window_size)
        if partition_by:
            expr = expr.over(partition_by)
    elif func in ("pct_change", "diff"):
        expr = getattr(base, func)()
        if partition_by:
            expr = getattr(base, func)().over(partition_by)
    elif func == "rank":
        expr = base.rank()
        if partition_by:
            expr = base.rank().over(partition_by)
    else:
        # cumsum, cumprod, cummax, cummin
        expr = getattr(base, func)()
        if partition_by:
            expr = getattr(base, func)().over(partition_by)

    result = lf.with_columns(expr.alias(out_col)).collect()
    abs_path = validate_path(file_path)
    state.set_df(abs_path, sheet_name, result)
    return (
        f"Added window column '{out_col}' using {func}.\n\n"
        + _df_to_markdown(result.head(20))
    )


def save_sheet(
    file_path: str,
    sheet_name: str,
    output_path: str,
    output_sheet: Optional[str] = None,
    overwrite_sheet: bool = True,
) -> str:
    """Save the current (possibly transformed) sheet data as a named sheet in an Excel workbook.

    If output_path already exists, the data is added as a new tab while all
    other sheets are preserved. If output_path does not exist, a new workbook
    is created. Use this to write transformed results back into a multi-sheet
    workbook alongside the original data.

    Args:
        file_path: Path to the source file.
        sheet_name: Sheet name to read from the session cache.
        output_path: Path to the target Excel workbook (.xlsx). Created if absent.
        output_sheet: Name for the sheet in the output workbook.
                      Defaults to sheet_name.
        overwrite_sheet: If the output sheet already exists, replace it (default True).
                         Set to False to raise an error instead.
    """
    import openpyxl

    df = _lazy(file_path, sheet_name).collect()
    abs_out = validate_path(output_path, write=True)
    out_sheet = output_sheet or sheet_name

    # Load or create the target workbook
    import os
    if os.path.exists(abs_out):
        wb = openpyxl.load_workbook(abs_out)
        if out_sheet in wb.sheetnames:
            if not overwrite_sheet:
                raise ValueError(
                    f"Sheet '{out_sheet}' already exists in '{abs_out}'. "
                    "Set overwrite_sheet=True to replace it."
                )
            del wb[out_sheet]
    else:
        wb = openpyxl.Workbook()
        # Remove the default empty sheet openpyxl creates
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet(out_sheet)

    # Write header row
    ws.append(df.columns)

    # Write data rows — use iter_rows for memory efficiency on large frames
    for row in df.iter_rows(named=False):
        ws.append(list(row))

    wb.save(abs_out)

    existing_sheets = wb.sheetnames
    return (
        f"Saved {df.height} rows × {df.width} columns as sheet '{out_sheet}' "
        f"in '{abs_out}'.\nWorkbook sheets: {existing_sheets}"
    )


def clone_sheet(
    file_path: str,
    sheet_name: str,
    new_sheet_name: str,
) -> str:
    """Copy a sheet under a new name within the same workbook and session cache.

    Use this before applying destructive transforms (filter, drop, cast) so
    you have a safe copy to fall back to. Both the original and the clone
    are available for subsequent tool calls.

    Args:
        file_path: Path to the Excel file (.xlsx only — openpyxl required).
        sheet_name: Name of the sheet to copy.
        new_sheet_name: Name for the new copy.
    """
    import os
    import openpyxl

    abs_path = validate_path(file_path)
    ext = os.path.splitext(abs_path)[1].lower()
    if ext not in (".xlsx", ".xlsm", ".xlsb"):
        raise ValueError(f"clone_sheet only supports .xlsx/.xlsm files, not '{ext}'.")

    wb = openpyxl.load_workbook(abs_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    if new_sheet_name in wb.sheetnames:
        raise ValueError(f"Sheet '{new_sheet_name}' already exists. Choose a different name.")

    # openpyxl copy_worksheet duplicates formatting + data
    wb.copy_worksheet(wb[sheet_name]).title = new_sheet_name
    wb.save(abs_path)

    # Also register the clone in the session cache so it's immediately usable
    original_df = _load(abs_path, sheet_name)
    state.set_df(abs_path, new_sheet_name, original_df.clone())

    return (
        f"Cloned '{sheet_name}' → '{new_sheet_name}' in '{abs_path}'.\n"
        f"Workbook sheets: {wb.sheetnames}"
    )


def export_data(
    file_path: str,
    sheet_name: str,
    output_path: str,
    format: Literal["csv", "xlsx", "parquet", "json"] = "csv",
) -> str:
    """Export the current (possibly transformed) sheet data to a file.

    Args:
        file_path: Path to the source file.
        sheet_name: Sheet name to export.
        output_path: Destination file path.
        format: Output format — "csv", "xlsx", "parquet", or "json".
    """
    # Collect lazily — only pulls what's in the plan
    df = _lazy(file_path, sheet_name).collect()
    abs_out = validate_path(output_path, write=True)
    if format == "csv":
        df.write_csv(abs_out)
    elif format == "xlsx":
        df.write_excel(abs_out)
    elif format == "parquet":
        df.write_parquet(abs_out)
    elif format == "json":
        df.write_ndjson(abs_out)
    else:
        raise ValueError(f"Unsupported format '{format}'. Use: csv, xlsx, parquet, json.")
    return f"Exported {df.height} rows × {df.width} columns to: {abs_out}"


def to_pandas(
    file_path: str,
    sheet_name: str,
    output_path: str,
    format: Literal["parquet", "csv"] = "parquet",
) -> dict:
    """Convert the current sheet data to a pandas-compatible file for use with ML libraries.

    Polars converts to pandas via Apache Arrow — zero-copy and dtype-preserving.
    Parquet is recommended because it retains all numeric types exactly (int32,
    float64, datetime, etc.) unlike CSV which loses type information. The returned
    snippet is ready to paste directly into a notebook or ML pipeline.

    Args:
        file_path: Path to the source file.
        sheet_name: Sheet name.
        output_path: Destination file path (.parquet or .csv).
        format: "parquet" (default, preserves dtypes) or "csv" (plain text).
    """
    try:
        import pyarrow  # noqa: F401 — required for Polars → pandas Arrow path
    except ImportError:
        raise ImportError(
            "pyarrow is required for Polars → pandas conversion. "
            "Install it with: pip install pyarrow"
        )

    df = _lazy(file_path, sheet_name).collect()
    abs_out = validate_path(output_path, write=True)

    if format == "parquet":
        df.write_parquet(abs_out)
        snippet = (
            f"import pandas as pd\n"
            f"df = pd.read_parquet('{abs_out}')\n"
            f"df.head()"
        )
    elif format == "csv":
        df.write_csv(abs_out)
        snippet = (
            f"import pandas as pd\n"
            f"df = pd.read_csv('{abs_out}')\n"
            f"df.head()"
        )
    else:
        raise ValueError(f"Unsupported format '{format}'. Use 'parquet' or 'csv'.")

    schema = {col: str(dtype) for col, dtype in zip(df.columns, df.dtypes)}

    return {
        "output_path": abs_out,
        "format": format,
        "rows": df.height,
        "columns": df.width,
        "schema": schema,
        "pandas_snippet": snippet,
    }


def run_query(
    file_path: str,
    sheet_name: str,
    sql: str,
    extra_sources: Optional[list[dict]] = None,
) -> str:
    """Run a SQL SELECT query against one or more sheets using Polars SQLContext.

    The primary sheet is always available as the table 'data'.
    Use extra_sources to register additional sheets under custom table names.

    For CSV and Parquet sources the query runs on a LazyFrame so filters,
    projections, and aggregations are pushed down into the file scan — only
    the rows and columns the query actually needs are read from disk.

    The full result is returned without a row cap — use LIMIT in your SQL to
    control output size when you only need a sample. For large datasets always
    aggregate in SQL rather than fetching raw rows.

    Examples:
        Aggregate:   SELECT Country, SUM(Revenue) as total FROM data GROUP BY Country ORDER BY total DESC
        Sample rows: SELECT * FROM data LIMIT 20
        Multi-sheet: SELECT a.*, b.Category FROM data a JOIN products b ON a.ProductId = b.Id

    Args:
        file_path: Path to the primary file.
        sheet_name: Primary sheet name (table alias: 'data').
        sql: Full SQL SELECT statement.
        extra_sources: Optional additional tables, e.g.
                       [{"file": "products.xlsx", "sheet": "Sheet1", "alias": "products"}]
    """
    tables: dict[str, pl.LazyFrame] = {"data": _lazy(file_path, sheet_name)}

    if extra_sources:
        for src in extra_sources:
            alias = src.get("alias")
            f = src.get("file", "")
            s = src.get("sheet", "")
            if not alias or not f or not s:
                raise ValueError("Each extra_source must have 'file', 'sheet', and 'alias' keys.")
            tables[alias] = _lazy(f, s)

    # eager=False keeps the query lazy until .collect(); Polars fuses all operations
    ctx = pl.SQLContext(tables, eager=False)
    result = ctx.execute(sql).collect()
    return f"Query returned {result.height} rows.\n\n" + _df_to_markdown(result, max_rows=result.height)
